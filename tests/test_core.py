from pathlib import Path
import io
import pandas as pd
from openpyxl import load_workbook
from config.settings import MAX_RISK_PER_TRADE,MIN_REQUIRED_RISK,MIN_RR_RATIO
from dashboard.dashboard_utils import build_single_sheet_master_excel
from papertrade.paper_trade_engine import PaperTradeEngine
from strategy.risk_engine import RiskEngine
from strategy.nifty500_price_action_strategies import evaluate_s1
ROOT=Path(__file__).resolve().parents[1]

def test_risk_engine_approves_target_risk_trade():
    result=RiskEngine().validate({"symbol":"TEST","signal":"BUY","entry":100.0,"stop_loss":98.0,"target":102.5},check_trade_count=False)
    assert result["approved"] is True;assert result["actual_risk"]==MAX_RISK_PER_TRADE;assert result["rr"]>=MIN_RR_RATIO;assert result["actual_risk"]>=MIN_REQUIRED_RISK

def test_risk_engine_rejects_wrong_side_stop():
    result=RiskEngine().validate({"symbol":"TEST","signal":"SELL","entry":100.0,"stop_loss":99.0,"target":98.0},check_trade_count=False)
    assert result["approved"] is False;assert any("SELL stop loss" in reason for reason in result["reasons"])

def test_clean_s1_requires_live_reclaim():
    g={"nifty500_change_pct":1.0,"sector_alignment_pct":1.0,"ad_ratio":2.0,"ad_coverage":500,"positive_sectors":10,"negative_sectors":5,"previous_candle_open":100,"previous_candle_close":101}
    assert evaluate_s1("TEST","BUY",110,100,90,99,115,111,**g) is not None
    assert evaluate_s1("TEST","BUY",110,100,90,99,115,109,**g) is None

def test_paper_pnl_buy_and_sell():
    assert PaperTradeEngine.calculate_pnl("BUY",100,102.5,10)==25.0;assert PaperTradeEngine.calculate_pnl("SELL",100,97.5,10)==25.0

def test_single_sheet_master_excel_contains_record_types():
    trades=pd.DataFrame([{"strategy":"S1","symbol":"ABC","signal":"BUY","entry":100,"pnl":25}]);signals=pd.DataFrame([{"strategy":"S2","symbol":"XYZ","signal":"SELL","approved":True}]);gaps=pd.DataFrame([{"Symbol":"LMN","TodayOpen":110,"PDH":100,"GapType":"GAP_UP"}])
    data=build_single_sheet_master_excel(trades,signals,gaps);workbook=load_workbook(io.BytesIO(data),read_only=True,data_only=True);assert workbook.sheetnames==["ALL DATA"]
    rows=list(workbook["ALL DATA"].iter_rows(values_only=True));header=rows[0];ri=header.index("Record Type");assert {r[ri] for r in rows[1:]}=={"TRADE","SIGNAL","GAP_BOARD"}

def test_dashboard_has_no_yahoo_or_legacy_strategy_imports():
    files=[ROOT/"main.py",ROOT/"engine"/"master_engine.py",ROOT/"market"/"price_data.py",ROOT/"strategy"/"nifty500_price_action_strategies.py"]
    for path in files:
        source=path.read_text(encoding="utf-8").lower();assert "yfinance" not in source;assert "open_reversal_engine" not in source;assert "gap_extension_reversal_engine" not in source
